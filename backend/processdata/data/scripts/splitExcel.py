import openpyxl
import re
from os import path


def splitExcel(excelPath):
    wb = openpyxl.load_workbook(excelPath)
    sheet = wb.active

    type_set = set()
    type2num_list = []
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue
        keys = row[1].value
        values = row[2].value
        keys = keys.split(";")
        values = values.split(";")

        type2num = {}
        for j in range(len(keys)):
            if keys[j] != "set":
                type_set.add(keys[j])
            type2num[keys[j]] = int(values[j])

        type2num_list.append(type2num)

    for i in range(len(type2num_list)):
        type2num = type2num_list[i]
        if "set" in type2num.keys():
            # type_set中每个制品都+set个数，没有的就新建
            for type in type_set:
                type2num[type] = (
                    type2num[type] + type2num["set"]
                    if type in type2num.keys()
                    else type2num["set"]
                )
        type2num.pop("set", None)
        # print(i + 2, type2num)

    type_list = list(type_set)
    type_list = sorted(
        type_list, key=lambda item: (bool(re.search("[a-zA-Z]", item)), item)
    )
    print(type_list)

    # 在excel的F-K行按type_list顺序写入type2num_list
    for i in range(len(type_list)):
        sheet.cell(row=1, column=6 + i).value = type_list[i]
        for j in range(len(type2num_list)):
            sheet.cell(row=2 + j, column=6 + i).value = (
                type2num_list[j][type_list[i]]
                if type_list[i] in type2num_list[j].keys()
                else 0
            )

    wb.save(excelPath)


p = path.dirname(__file__) + "/../test_excel/" + "深海之灵.xlsx"
splitExcel(p)
