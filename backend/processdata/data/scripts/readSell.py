import openpyxl
import re
from os import path
import json
import warnings
from openpyxl import Workbook  # Ensure openpyxl is imported

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# 解析cn和qq在同一列的函数（已废弃）
def processString(input_string):
    # 初始化返回变量
    cn = ""
    qq = ""

    # 搜索 cn+群内qq: 模式
    pattern = r"cn\+群内qq:(.*)"
    match = re.search(pattern, input_string)

    if not match:  # 如果没有 cn+群内qq:，则直接返回
        return cn, qq

    cn_qq_str = match.group(1)

    # 特殊情况1：cn和qq之间只用一个特殊字符分隔
    separators = r"[\+\＋\-\－\—\,\，\s\;\；]"
    simple_pattern = rf"^(.+?){separators}(\d+)$"
    match = re.match(simple_pattern, cn_qq_str)
    if match:
        return match.group(1), match.group(2)

    # 特殊情况2，cn也全是数字
    digit_pattern = rf"^(\d+){separators}(\d+)$"
    match = re.match(digit_pattern, cn_qq_str)
    # 直接返回cn和qq全是数字的情况
    if match:
        return match.group(1), match.group(2)

    # 尝试从字符串中提取 cn 和 qq
    # 提取 cn
    cn_pattern = r"(.*?)\s*(?=\d{6,})|(.+)$"
    matches = re.findall(cn_pattern, cn_qq_str)

    if matches:
        # 如果第一个捕获组为空（没有匹配到后面跟随至少6个数字的内容）
        # 则使用第二个捕获组的值（即整个cn_qq_str的值）
        # 适用于只写了cn没有qq的情况
        cn = matches[0][0] or matches[0][1]
        cn = cn.strip()  # 移除两端的空白字符
        if cn.endswith(("+", "＋", "-", "－", "—", ",", "，", ";", "；")):
            cn = cn[:-1]

    # 如果没有提取到 cn
    if not cn:
        cn = "??????"

    # 提取 qq
    qq_pattern = r"\d{6,}"
    match = re.search(qq_pattern, cn_qq_str)

    if match:
        qq = match.group()
    else:
        # 只写了qq没有cn的情况
        qq = "??????"

    return cn, qq


def processRow_old(row, single_sheet_data):
    cn = ""
    qq = ""
    num = 0
    # 单种谷子的单行的数据
    row_data = []
    for i in range(len(row)):
        # 读取第一列，处理谷子和qq
        if i == 0:
            cn, qq = processString(row[i].value)

            # 标题行
            if cn == "" and qq == "":
                row_data.append(row[i].value)
            # cn，qq行
            else:
                row_data.extend([cn, qq])
        # 读取第二列，处理数量
        if i == 1 and row[i].value is not None and isinstance(row[i].value, int):
            num = int(row[i].value)
            row_data.append(num)
        elif i == 1 and isinstance(row[i].value, str):
            # 可能是“数量”这两个字（在第一行），是中文的话直接插入
            pattern = re.compile(r"[\u4e00-\u9fff]+")
            is_chinese = bool(pattern.search(row[i].value))
            if is_chinese:
                row_data.append(row[i].value)
            # 不是的话就是可能为字符串形式的数字，转换为int
            else:
                row_data.append(int(row[i].value))
        elif i == 1 and row[i].value is None:
            return
        # 读取第三列，处理状态
        if i == 2:
            if row[i].value is None:
                row_data.append("none")
            else:
                row_data.append(row[i].value)

    single_sheet_data.append(row_data)


def processSheetSingleType(sheet, sheetdatas):
    # 单种谷子的数据
    single_sheet_data = []
    # 读取整个工作表的数据
    for cnt, row in enumerate(sheet.iter_rows()):
        row_data = []
        # 如果到达某一行为空，则遍历完毕
        if row[0].value is None:
            break

        # 处理标题行
        if cnt == 0:
            title = row[0].value
            row_data.append(title)

            single_sheet_data.append(row_data)
            continue
        # 跳过第二行（cn qq 数量 状态）
        elif cnt == 1:
            continue

        # 数据行
        cn = row[0].value
        qq = row[1].value
        num = row[2].value
        status = row[3].value if row[3].value is not None else "none"

        single_sheet_data.append([cn, qq, num, status])

    # print()
    # print(single_sheet_data)
    # 插入单个子表数据
    sheetdatas.extend(single_sheet_data)


def processSheetMultiType(sheet, condition_index, sheetdatas):
    # 表示要循环几次，即有一对多的情况下有几个角色
    count = condition_index - 3  # 比如5-3=2(3代表cn，qq，状态)

    for character_index in range(count):
        # 单种谷子的数据
        single_sheet_data = []
        card_name = None
        character_name = None
        # 读取整个工作表的数据
        for line_cnt, row in enumerate(sheet.iter_rows()):
            # 如果到达某一行为空，则遍历完毕
            if row[0].value is None:
                break

            # 第一行，保存谷子名，跳过
            if line_cnt == 0:
                card_name = row[0].value
                # print(card_name)
                continue
            # 第二行，保存角色名
            elif line_cnt == 1:
                character_name = row[character_index + 2].value
                # print(character_name)
                continue

            # 数据行(cn,qq,数量)
            cn = row[0].value
            qq = row[1].value
            num = row[character_index + 2].value
            status = (
                row[condition_index - 1].value
                if row[condition_index - 1].value is not None
                else "none"
            )

            # 跳过数量为空的行，说明这个cn没有买这个type的谷子
            if num is None:
                continue

            single_sheet_data.append([cn, qq, num, status])

        # 处理谷子名
        # 把角色名和谷子名加起来，再按照顺序加上card_id（19-1,19-2,19-3...）
        match = re.match(r"\d+", card_name)
        card_id_end_pos = match.end()

        card_name = (
            card_name[:card_id_end_pos]
            + "_"
            + str(character_index + 1)
            + card_name[card_id_end_pos:]
        )
        card_name += "-" + character_name

        single_sheet_data.insert(0, [card_name])

        # print()
        # print(single_sheet_data)

        # 插入单个子表（角色）数据
        sheetdatas.extend(single_sheet_data)
    # print(sheetdatas)


def readSellInfo(path):
    # p = path.dirname(__file__) + "/../test_excel/" + excel_name
    # 读取Excel文件
    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_names = wb.sheetnames[3:]
    # print(sheet_names)

    # 一个表格的所有谷子信息
    # 每一个元素对应一种谷子的信息
    sheetdatas = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]  # 修改为实际的工作表名

        # 获取状态所在的列
        condition_index = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "状态":
                    condition_index = cell.col_idx
                    break

        # 一个谷子对应多个角色的情况
        # 一行大于三列：cn,qq，角色1，角色2....角色N，数量，状态
        if condition_index > 4:
            processSheetMultiType(sheet, condition_index, sheetdatas)

        # 一行等于三列：cn,qq，数量，状态
        elif condition_index == 4:
            processSheetSingleType(sheet, sheetdatas)

    # 关闭Excel文件
    wb.close()
    return sheetdatas


def writeJsonFile(json_name, excel_data):
    # 分割每个谷子的数据，然后存入字典
    split_points = []
    for i in range(len(excel_data)):
        if len(excel_data[i]) == 1:
            # print(excel_data[i])
            split_points.append(i)

    # 在 split_points 列表末尾添加 excel_data 的长度
    # 为了能够在下面的循环中把最后一个谷子的数据也加入字典
    split_points.append(len(excel_data))
    dict_data = {}
    for i in range(len(split_points) - 1):
        split_index = split_points[i]
        id_and_name = excel_data[split_index][0]
        # print(id_and_name)

        # 如果长度为1，说明是标题行，获取了card_name和card_id就进行下一次循环
        match1 = re.search(r"\d+", id_and_name)
        # 一个谷子多个角色的情况
        match2 = re.search(r"\d+\_\d+", id_and_name)
        if match1 and match2 is None:
            card_id = match1.group()

            dict_data[card_id] = excel_data[split_index : split_points[i + 1]]
        elif match2:
            card_id = match2.group()

            dict_data[card_id] = excel_data[split_index : split_points[i + 1]]
    # print(dict_data)

    p = path.dirname(__file__) + "/../json/" + json_name

    # Convert list to JSON and save it to the file
    with open(p, "w", encoding="utf-8") as json_file:
        json.dump(dict_data, json_file, ensure_ascii=False, indent=4)


def ReadSell(file_path):
    excel_data = readSellInfo(file_path)
    writeJsonFile("selldata.json", excel_data)


if __name__ == "__main__":
    file_name = (
        path.dirname(__file__)
        + "/../test_excel/selldata/"
        + "selldata_2023_08_29_45.xlsx"
    )

    excel_data = readSellInfo(file_name)

    writeJsonFile("selldata.json", excel_data)
