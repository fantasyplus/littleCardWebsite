import openpyxl
from os import path
import json


def readCardInfo(path):
    # p = path.dirname(__file__) + "/../excel/" + filename
    # 打开Excel文件
    workbook = openpyxl.load_workbook(path)
    sheet_names = workbook.sheetnames
    # 选择工作表
    sheet = workbook[sheet_names[0]]

    # 存储数据的字符串序列
    data = []

    # 遍历每一行
    for row in sheet.iter_rows(values_only=True):
        # 将每个单元格的数据转换为字符串并添加到data列表中
        row_data = [str(cell) for cell in row]
        data.append(row_data)

    return data

def ReadCard(file_path):
    excel_data = readCardInfo(file_path)

    file_path = "carddata.json"
    p = path.dirname(__file__) + "/../json/" + file_path
    data = {"title": excel_data[1], "data": excel_data[2:]}
    # Convert list to JSON and save it to the file
    with open(p, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4)

